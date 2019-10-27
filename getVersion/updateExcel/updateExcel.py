#!/usr/bin/env python

import json
import os
import sys
import argparse
import configparser
import six
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

config_file = 'xlsx_inventory.cfg'
default_group = 'NO_GROUP'

try:
    FileNotFoundError
except NameError:
    FileNotFoundError = IOError

def find_config_file():
    env_name = 'EXCEL_INVENTORY_CONFIG'
    if env_name in os.environ:
        return os.environ[env_name]
    else:
        return config_file


def main():
    config_path = find_config_file()
    config = load_config(config_path)
    file_location = config['xlsx_inventory_file']

    try:
        wb = load_workbook(file_location)
        if 'sheet' in config:
            sheet = wb[config['sheet']]
        else:
            sheet = wb.active
        if os:
            update_excel(sheet=sheet, brand=str(sys.argv[1]), host_ip=str(sys.argv[2]),
                     host_name=str(sys.argv[3]), sw_version=str(sys.argv[4]), os=str(sys.argv[5]))
        else:
            update_excel(sheet=sheet, brand=str(sys.argv[1]), host_ip=str(sys.argv[2]),
                     host_name=str(sys.argv[3]), sw_version=str(sys.argv[4]), os = " ")

        wb.save(file_location)

    except FileNotFoundError as e:
        print(
            '\033[91mFile Not Found! Check %s configuration file!'
            ' Is the `xlsx_inventory_file` path setting correct?\033[0m' % config_path)
        print(e)
        exit(1)
    except KeyError as e:
        print(
            '\033[91mKey Error! Check %s configuration file! Is the `sheet` name setting correct?\033[0m' % config_path)
        print(e)
        exit(1)

    except zipfile.BadZipfile as e:
        print("Excel is opened by other process. Restarting main()")
        main()
        exit(0)


def create_col(sheet, condition):
    print("Adding column for %s" %(condition))
    if "ansible_net_hostname" is condition:
        max_col = 0
    else:
        max_col = sheet.max_column
    sheet.insert_cols(max_col + 1)
    sheet.cell(row=1, column=max_col + 1).value = condition
    print max_col
    return max_col


def find_column(sheet, condition):
    for col in sheet.columns:
        if condition == col[0].value:
            return column_index_from_string(coordinate_from_string(col[0].column + '1')[0]) - 1
    return create_col(sheet, condition)


def update_excel(sheet, brand, host_ip, host_name, sw_version, os ):
    hostname_col = find_column(sheet, "ansible_net_hostname")
    ip_col = find_column(sheet, "ansible_host")
    brand_col = find_column(sheet, "brand")
    sw_col = find_column(sheet, "OS version")
    os_col = find_column(sheet, "ansible_network_os")

    print ip_col
    print hostname_col
    print brand_col
    print sw_col
    print os_col

    for row in sheet.rows:
        if host_ip == row[ip_col].value:
            row[hostname_col].value = host_name
            row[sw_col].value = sw_version
            row[brand_col].value = brand
            row[os_col].value = os

def load_config(config_path):
    config = configparser.ConfigParser()
    config['DEFAULT'] = {'hostname_col': 'A', 'group_by_col': 'B'}
    if len(config.read(config_path)) > 0:
        return config['xlsx_inventory']
    else:
        print('\033[91mConfiguration File "%s" not Found!\033[0m' % config_path)
        exit(1)


if __name__ == "__main__":
    main()
